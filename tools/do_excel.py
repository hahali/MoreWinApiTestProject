#import xlrd
import openpyxl
from tools import project_path
from tools import read_conf
import logging
import logging.config

CON_LOG=project_path.log_path #获取日志配置文件
logging.config.fileConfig(CON_LOG) # 采用配置文件配置日志
logging=logging.getLogger("测试") #

class DoExcel:
    #读取excel数据
    def get_data(self,filename):
        #打开excel文件
        wb=openpyxl.load_workbook(filename)
        #获取login工作表
        mode=eval(read_conf.ReadConfig.read_config(project_path.case_config_path, 'MODE', 'mode'))
        test_data = []#创建空列表，用于存储数据
        for key in mode: #遍历存在配置文件的字段
            sheet=wb[key]
            #获取所有的有效行数
            nrows=sheet.max_row
            logging.info(nrows)
            if mode[key]=='all':
                for i in range(2,nrows+1):
                    # 创建空字典，用于赋值
                    rownum_data={}
                    rownum_data["id"] = sheet.cell(i, 1).value
                    rownum_data["method"] = sheet.cell(i, 2).value
                    rownum_data["url"] = sheet.cell(i,3).value
                    rownum_data["header"] = sheet.cell(i, 4).value
                    #rownum_data["data"] = sheet.cell(i, 5).value

                    if sheet.cell(i,5).value.find("${iphone}")!=-1: #查找从excel读取的data是否有${iphone}字符串
                        tel = self.get_tel(filename, "init")  # 从excel中动态获取注册的手机号
                        rownum_data["data"]=sheet.cell(i, 5).value.replace("${iphone}",str(tel)) #把${iphone}替换成excel中获取的手机号码
                        #logging.info('tel数据类型：{0}'.format(type(tel))) #查看读取的tel类型
                        logging.info('tel数据类型：{0}'.format(type(tel)))
                        tel=tel+1
                        self.write_back_tel(filename,"init",tel)
                    else:
                        rownum_data["data"] = sheet.cell(i, 5).value
                    logging.info(type(rownum_data["data"]))
                    rownum_data["tital"] = sheet.cell(i, 6).value
                    rownum_data["expect"] = sheet.cell(i, 7).value
                    #logging.info(type(rownum_data["expect"]))
                    rownum_data["sheet_name"] = key
                    test_data.append(rownum_data)
            else:
                for case_id in mode[key]:
                    # 创建空字典，用于赋值
                    rownum_data={}
                    rownum_data["id"] = sheet.cell(case_id+1, 1).value
                    rownum_data["method"] = sheet.cell(case_id+1, 2).value
                    rownum_data["url"] = sheet.cell(case_id+1,3).value
                    rownum_data["header"] = sheet.cell(case_id+1, 4).value
                    #rownum_data["data"] = sheet.cell(case_id+1, 5).value

                    if sheet.cell(case_id+1,5).value.find("${iphone}")!=-1: #查找从excel读取的data是否有${iphone}字符串
                        tel = self.get_tel(filename, "init")  # 从excel中动态获取注册的手机号
                        rownum_data["data"]=sheet.cell(case_id+1, 5).value.replace("${iphone}",str(tel)) #把${iphone}替换成excel中获取的手机号码
                        logging.info('tel数据类型：{0}'.format(type(tel))) #查看读取的tel类型
                        tel=tel+1
                        self.write_back_tel(filename,"init",tel)
                    else:
                        rownum_data["data"] = sheet.cell(case_id+1, 5).value
                    logging.info(type(rownum_data["data"]))
                    rownum_data["tital"] = sheet.cell(case_id+1, 6).value
                    rownum_data["expect"] = sheet.cell(case_id+1, 7).value
                    #logging.info(type(rownum_data["expect"]))
                    rownum_data["sheet_name"] = key
                    test_data.append(rownum_data)

        logging.info("读取excel用例数据返回值：{0}".format(test_data))
        return test_data


    #把接口返回结果写入excel
    def write_data(self,filename,sheel_name,i,value,result):
        wb=openpyxl.load_workbook(filename)
        sheel=wb[sheel_name]
        sheel.cell(i,8).value=value
        sheel.cell(i,9).value=result
        wb.save(filename)

    #获取注册的手机号码
    def get_tel(self,filename,sheel_name):
        wb= openpyxl.load_workbook(filename)
        sheel=wb[sheel_name]
        tel=sheel.cell(1,1).value
        return tel

    #把新的手机号码写回excel
    def write_back_tel(self,filename,sheel_name,tel):
        wb=openpyxl.load_workbook(filename)
        sheel=wb[sheel_name]
        sheel.cell(1,1).value=tel
        wb.save(filename)

    #获取域名地址
    def get_url(self,filename,sheel_name):
        wb= openpyxl.load_workbook(filename)
        sheel=wb[sheel_name]
        url=sheel.cell(3,1).value
        return url



if __name__=="__main__":

    data=DoExcel().get_data('../testData/KXliveApiCase.xlsx')
    #logging.info(data)
    tel=DoExcel().get_tel('../testData/KXliveApiCase.xlsx','init')
    logging.info(tel)
    logging.info(type(tel))